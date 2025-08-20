"""Excel tracking utilities for Resume Automation System"""

import os
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


class ResumeTracker:
    """Handles tracking resume applications in an Excel spreadsheet"""
    
    REQUIRED_COLUMNS = [
        'Company',
        'Department', 
        'Role',
        'Salary',
        'Application Date',
        'Application Page'
    ]
    
    def __init__(self, worksheet_path: str):
        """
        Initialize resume tracker
        
        Args:
            worksheet_path: Path to the Excel file (relative or absolute)
        """
        self.worksheet_path = Path(worksheet_path)
        
    def ensure_worksheet_exists(self) -> bool:
        """
        Ensure the Excel worksheet exists with proper headers
        
        Returns:
            bool: True if worksheet was created/verified successfully
        """
        try:
            # Create directory if it doesn't exist
            self.worksheet_path.parent.mkdir(parents=True, exist_ok=True)
            
            if self.worksheet_path.exists() and self.worksheet_path.stat().st_size > 0:
                # Verify existing file has correct headers
                return self._verify_headers()
            else:
                # Create new worksheet
                return self._create_new_worksheet()
                
        except Exception as e:
            print(f"Error ensuring worksheet exists: {str(e)}")
            return False
    
    def _create_new_worksheet(self) -> bool:
        """Create a new Excel worksheet with proper formatting"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Resume Applications"
            
            # Add headers
            for col_idx, header in enumerate(self.REQUIRED_COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                # Header formatting
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Set column widths
            column_widths = {
                'A': 20,  # Company
                'B': 15,  # Department
                'C': 25,  # Role
                'D': 12,  # Salary
                'E': 15,  # Application Date
                'F': 40   # Application Page (URL)
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Save workbook
            wb.save(self.worksheet_path)
            print(f"Created new resume tracking worksheet: {self.worksheet_path}")
            return True
            
        except Exception as e:
            print(f"Error creating worksheet: {str(e)}")
            return False
    
    def _verify_headers(self) -> bool:
        """Verify existing worksheet has correct headers"""
        try:
            # Read first row to check headers
            df = pd.read_excel(self.worksheet_path, nrows=1, engine='openpyxl')
            existing_columns = list(df.columns)
            
            # Check if all required columns exist
            missing_columns = [col for col in self.REQUIRED_COLUMNS if col not in existing_columns]
            
            if missing_columns:
                print(f"Warning: Worksheet missing columns: {missing_columns}")
                # Automatically add missing columns
                return self._add_missing_columns(missing_columns)
                
            print(f"Resume tracking worksheet verified: {self.worksheet_path}")
            return True
            
        except Exception as e:
            print(f"Error verifying headers: {str(e)}")
            return False
    
    def _add_missing_columns(self, missing_columns: list) -> bool:
        """Add missing columns to existing worksheet"""
        try:
            # Load existing data
            df = pd.read_excel(self.worksheet_path, engine='openpyxl')
            
            # Add missing columns with empty values
            for col in missing_columns:
                df[col] = ""
            
            # Reorder columns to match REQUIRED_COLUMNS
            df = df.reindex(columns=self.REQUIRED_COLUMNS)
            
            # Save updated worksheet
            self._save_formatted_excel(df)
            
            print(f"Successfully added missing columns: {missing_columns}")
            return True
            
        except Exception as e:
            print(f"Error adding missing columns: {str(e)}")
            return False
    
    def add_application_record(self, 
                             company: str,
                             department: str,
                             role: str,
                             salary: str = "",
                             application_date: Optional[datetime] = None,
                             application_page: str = "") -> bool:
        """
        Add a new application record to the tracking worksheet
        
        Args:
            company: Company name
            department: Department or team name
            role: Job title/position
            salary: Expected or posted salary (optional)
            application_date: Date of application (defaults to today)
            application_page: Job description URL (optional)
            
        Returns:
            bool: True if record was added successfully
        """
        try:
            # Ensure worksheet exists
            if not self.ensure_worksheet_exists():
                return False
            
            # Set default application date
            if application_date is None:
                application_date = datetime.now()
            
            # Format application date
            formatted_date = application_date.strftime("%Y-%m-%d")
            
            # Create record data
            new_record = {
                'Company': company,
                'Department': department,
                'Role': role,
                'Salary': salary,
                'Application Date': formatted_date,
                'Application Page': application_page
            }
            
            # Check if this exact record already exists (to avoid duplicates)
            if self._record_exists(company, role, formatted_date):
                print(f"Record already exists for {company} - {role} on {formatted_date}")
                return True
            
            # Load existing data
            try:
                df = pd.read_excel(self.worksheet_path, engine='openpyxl')
            except Exception:
                # If file is corrupted or empty, create empty DataFrame
                df = pd.DataFrame(columns=self.REQUIRED_COLUMNS)
            
            # Add new record
            new_row = pd.DataFrame([new_record])
            df = pd.concat([df, new_row], ignore_index=True)
            
            # Sort by application date (newest first)
            df['Application Date'] = pd.to_datetime(df['Application Date'])
            df = df.sort_values('Application Date', ascending=False)
            df['Application Date'] = df['Application Date'].dt.strftime('%Y-%m-%d')
            
            # Save back to Excel with formatting
            self._save_formatted_excel(df)
            
            print(f"Added application record: {company} - {role}")
            return True
            
        except Exception as e:
            print(f"Error adding application record: {str(e)}")
            return False
    
    def _record_exists(self, company: str, role: str, application_date: str) -> bool:
        """Check if a record with the same company, role, and date already exists"""
        try:
            df = pd.read_excel(self.worksheet_path, engine='openpyxl')
            matching_records = df[
                (df['Company'].str.lower() == company.lower()) & 
                (df['Role'].str.lower() == role.lower()) &
                (df['Application Date'] == application_date)
            ]
            return len(matching_records) > 0
        except Exception:
            return False
    
    def check_duplicate_resume(self, company: str, role: str, application_page: str = "") -> bool:
        """
        Check if a resume already exists for this company/role/URL combination
        
        Args:
            company: Company name
            role: Job title/position  
            application_page: Job description URL (optional)
            
        Returns:
            bool: True if duplicate exists, False if no duplicate found
        """
        try:
            # Ensure worksheet exists first
            if not self.ensure_worksheet_exists():
                return False
                
            df = pd.read_excel(self.worksheet_path, engine='openpyxl')
            
            if len(df) == 0:
                return False  # No records exist yet
            
            # Check for exact company + role match
            company_role_matches = df[
                (df['Company'].str.lower() == company.lower()) & 
                (df['Role'].str.lower() == role.lower())
            ]
            
            # If we have an application page URL, check for exact URL match
            if application_page and 'Application Page' in df.columns:
                url_matches = company_role_matches[
                    company_role_matches['Application Page'].str.lower() == application_page.lower()
                ]
                if len(url_matches) > 0:
                    return True  # Found exact company + role + URL match
            
            # If no URL provided or no URL match, check for company + role match from today
            today = datetime.now().strftime("%Y-%m-%d")
            recent_matches = company_role_matches[
                company_role_matches['Application Date'] == today
            ]
            
            return len(recent_matches) > 0
            
        except Exception as e:
            print(f"Error checking for duplicate resume: {str(e)}")
            return False  # If error occurs, allow generation to proceed
    
    def _save_formatted_excel(self, df: pd.DataFrame):
        """Save DataFrame to Excel with proper formatting"""
        try:
            # Load existing workbook to preserve formatting
            if self.worksheet_path.exists():
                wb = load_workbook(self.worksheet_path)
                ws = wb.active
                
                # Clear existing data (keep headers)
                ws.delete_rows(2, ws.max_row)
                
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Resume Applications"
            
            # Write data starting from row 2 (after headers)
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Ensure headers are formatted properly
            for col_idx, header in enumerate(self.REQUIRED_COLUMNS, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Set column widths
            column_widths = [20, 15, 25, 12, 15, 40]  # Company, Department, Role, Salary, Date, Application Page
            for col_idx, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col_idx)].width = width
            
            # Save workbook
            wb.save(self.worksheet_path)
            
        except Exception as e:
            print(f"Error saving formatted Excel: {str(e)}")
            # Fallback to simple DataFrame save
            df.to_excel(self.worksheet_path, index=False, engine='openpyxl')
    
    def get_application_stats(self) -> Dict[str, Any]:
        """Get statistics about applications"""
        try:
            df = pd.read_excel(self.worksheet_path, engine='openpyxl')
            
            total_applications = len(df)
            companies_applied = df['Company'].nunique()
            roles_applied = df['Role'].nunique()
            
            # Applications this month
            df['Application Date'] = pd.to_datetime(df['Application Date'])
            current_month = datetime.now().replace(day=1)
            this_month_applications = len(df[df['Application Date'] >= current_month])
            
            return {
                'total_applications': total_applications,
                'companies_applied': companies_applied,
                'roles_applied': roles_applied,
                'this_month_applications': this_month_applications,
                'worksheet_path': str(self.worksheet_path)
            }
            
        except Exception as e:
            print(f"Error getting application stats: {str(e)}")
            return {
                'total_applications': 0,
                'companies_applied': 0,
                'roles_applied': 0,
                'this_month_applications': 0,
                'worksheet_path': str(self.worksheet_path)
            }


def create_resume_tracker(config: dict) -> Optional[ResumeTracker]:
    """
    Factory function to create a ResumeTracker from configuration
    
    Args:
        config: Configuration dictionary with file_organization settings
        
    Returns:
        ResumeTracker instance or None if disabled
    """
    file_config = config.get('file_organization', {})
    
    # Check if resume tracking is enabled
    if not file_config.get('enable_resume_tracking', False):
        return None
    
    # Get worksheet path
    worksheet_path = file_config.get('resume_tracking_worksheet')
    if not worksheet_path:
        print("Warning: resume_tracking_worksheet not configured")
        return None
    
    # Make path absolute if it's relative
    if not os.path.isabs(worksheet_path):
        drive_root = file_config.get('drive_for_mac_root', '')
        if drive_root:
            worksheet_path = os.path.join(drive_root, worksheet_path)
        else:
            print("Warning: Cannot resolve relative worksheet path without drive_for_mac_root")
            return None
    
    return ResumeTracker(worksheet_path)